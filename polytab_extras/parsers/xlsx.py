#
# xlsx.py
#

from polytab.parsers import base
import openpyxl


class XLSXParser(base.Parser):
    READ_MODE = 'rb'
    WRITE_MODE = 'wb'

    def _set_argparser_options(self):
        """
        Creates an ArgumentParser with the parser's allowed arguments.
        """
        super(XLSXParser, self)._set_argparser_options()

        self._inparser.add_argument('--infile-sheet-name',
            nargs='?',
            dest='sheet_name')
        self._inparser.add_argument('--infile-sheet-index',
            type=int,
            default=1,
            dest='sheet_id')
        self._inparser.add_argument('--infile-dim',
            dest='dimension')
        self._inparser.add_argument('--infile-no-header',
            action='store_false',
            dest='hasheader')
        self._inparser.add_argument('--infile-start-row',
            type=int,
            dest='start_row')
        self._inparser.add_argument('--infile-end-row',
            type=int,
            dest='end_row')

    def read(self, fileobj):
        """
        Read an xlsx file
        :param fileobj: Open file handle
        """
        wbook = openpyxl.load_workbook(fileobj, read_only=True)

        if self.sheet_id is not None:
            if 0 < self.sheet_id <= len(wbook.sheetnames):
                wsheet = wbook[wbook.sheetnames[self.sheet_id - 1]]
            else:
                raise InvalidSheetError(self.sheet_id, 'index', options=wbook.sheetnames)
        else:
            if self.sheet_name in wbook:
                wsheet = wbook[self.sheet_name]
            else:
                raise InvalidSheetError(self.sheet_name, 'name', options=wbook.sheetnames)

        rows = []
        for row in wsheet.rows:
            rows.append([x.value for x in row])

        header = rows.pop(0)

        return header, rows

    def write(self, fileobj):
        """
        Write an xlsx file
        :param fileobj: Open file handle
        """
        wbook = openpyxl.Workbook(write_only=True)
        wsheet = wbook.create_sheet()

        if self.header is not None: # and self.hasheader is not False:
            wsheet.append(self.header)

        for row in self.rows:
            wsheet.append(row)

        wbook.save(fileobj)


class InvalidSheetError(Exception):
    MESSAGE = "Sheet {} '{}' is not in workbook. Please select " \
        "one of the following: {}"

    def __init__(self, sheet, identifier, options=None):
        self.sheet = sheet
        self.identifier = identifier
        self.options = ', '.join(options) if options is not None else ''

    def __str__(self):
        return self.MESSAGE.format(self.identifier, self.sheet, self.options)
